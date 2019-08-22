#!/usr/bin/env python3.5
#
import pandas
from openpyxl import Workbook
import datetime

class Report(object):
    def __init__(self):
        #self.data = {"分辨率": [], "发送端丢包设置": [], "发送端编码帧率(F/S)": [], "发送端Qos之前丢包率":[], "发送端Qos前占用带宽(kbps)": [], "发送端Qos后丢包率": [], "发送端Qos后占用带宽(kbps)": [], "测试时长（min）": [], "接收端解码帧率(F/S)": [], "最大连续丢包个数": [], "nack平均补偿码率(bps)": [], "主观感受": []}
        #pd = pandas.DataFrame(self.data)
        #excel = pd.to_excel('report.xlsx')
        
        self.vedio_name_list = ["分辨率", "发送端丢包设置",  "发送端编码帧率(F/S)", "发送端Qos之前丢包率", "发送端Qos前占用带宽(kbps)", "发送端Qos后丢包率", "发送端Qos后占用带宽(kbps)", "测试时长(min)", "接收端解码帧率(F/S)", "最大连续丢包个数", "nack平均补偿码率(bps)", "主观感受"]
        self.audio_name_list = ["发送端丢包率设置", "发送端Qos前丢包率", "发送端Qos前占用带宽(Kbps)", "发送端Qos后丢包率", "发送端Qos后占用带宽(Kbps)", "最大连续丢包个数", "nack平均补偿码率(Kbps)", "主观感受"]
        wb = Workbook()
        ws1 = wb.create_sheet("视频抗丢包", 0)
        ws2 = wb.create_sheet("音频抗丢包", 1)
        """
        index = 0
        for cell in ws1[1:len(self.col_name_list)]:
            print(cell)
            cell[0].value = self.col_name_list[index]
            index += 1
        """
        ws1.append(self.vedio_name_list)
        ws2.append(self.audio_name_list)
        wb.save('tt.xlsx')

test = Report()
